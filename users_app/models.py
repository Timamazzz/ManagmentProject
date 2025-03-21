import calendar
from calendar import monthrange
from datetime import date, datetime

import openpyxl
from django.contrib.auth.models import AbstractUser
from django.core.files.base import ContentFile
from django.db import models
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from django_jsonform.models.fields import JSONField
from django.db import transaction

from users_app.report_utils import get_volunteers_for_report, get_worked_days


# Create your models here.
class User(AbstractUser):
    pass


class Volunteer(models.Model):
    STATUS_CHOICES = [
        ('active', 'Активен'),
        ('dismissed', 'Уволен'),
        ('reserve', 'Резерв'),
    ]
    number_service = models.CharField(max_length=256, verbose_name="Личный  номер")

    first_name = models.CharField(max_length=30, verbose_name="Имя", null=True)
    last_name = models.CharField(max_length=30, verbose_name="Фамилия", null=True)
    patronymic = models.CharField(max_length=30, blank=True, null=True, verbose_name="Отчество")

    birthday = models.DateField(verbose_name="Дата рождения", null=True)

    passport_series = models.CharField(max_length=10, verbose_name="Серия паспорта", null=True)
    passport_number = models.CharField(max_length=15, verbose_name="Номер паспорта", null=True)
    passport_issued = models.CharField(max_length=255, verbose_name="Кем выдан паспорт", null=True)
    passport_issue_date = models.DateField(verbose_name="Дата выдачи паспорта", null=True)

    contract_date = models.DateField(verbose_name="Дата заключения контракта", null=True)
    order_number = models.CharField(max_length=50, verbose_name="Номер приказа", null=True)
    enrollment_date = models.DateField(verbose_name="Дата зачисления в добровольческое формирование", null=True)

    salary_amount = models.DecimalField(max_digits=10, decimal_places=2, verbose_name="Размер денежной выплаты",
                                        default=0.00, null=True, blank=True)

    bic = models.CharField(max_length=9, verbose_name="БИК", null=True, blank=True)
    bank_name = models.CharField(max_length=255, verbose_name="Наименование кредитной организации", null=True,
                                 blank=True)
    correspondent_account = models.CharField(max_length=20, verbose_name="Корреспондентский счет", null=True,
                                             blank=True)
    checking_account = models.CharField(max_length=20, verbose_name="Расчетный счет", null=True, blank=True)
    inn = models.CharField(max_length=12, verbose_name="ИНН", null=True, blank=True)
    kpp = models.CharField(max_length=9, verbose_name="КПП", null=True, blank=True)

    dismissal_date = models.DateField(blank=True, null=True, verbose_name="Дата увольнения")
    dismissal_order_number = models.CharField(max_length=50, blank=True, null=True,
                                              verbose_name="Номер приказа об увольнении")

    rank = models.CharField(max_length=1024, verbose_name="Звание", blank=True, null=True)
    status = models.CharField(max_length=10, choices=STATUS_CHOICES, default='active', verbose_name="Статус")

    salary = models.PositiveIntegerField("Оклад", default=0, null=True, blank=True)

    class Meta:
        verbose_name = "Доброволец"
        verbose_name_plural = "Добровольцы"
        permissions = [
            ("can_manage_reserve", "Может управлять резервистами"),
        ]

    def __str__(self):
        return f"{self.last_name} {self.first_name} ({self.number_service})"

    def clean(self):
        """Пользовательская валидация для добровольца"""
        if self.status == 'dismissed':
            if not self.dismissal_date:
                raise ValidationError({
                    'dismissal_date': _("При увольнении необходимо указать дату увольнения!")
                })
            if not self.dismissal_order_number:
                raise ValidationError({
                    'dismissal_order_number': _("При увольнении необходимо указать номер приказа!")
                })
            if self.dismissal_date and self.enrollment_date and self.dismissal_date < self.enrollment_date:
                raise ValidationError({
                    'dismissal_date': _("Дата увольнения не может быть раньше даты зачисления!")
                })

    def save(self, *args, **kwargs):
        self.full_clean()
        super().save(*args, **kwargs)


class Combat(models.Model):
    """Модель боевых выплат"""
    volunteer = models.ForeignKey(Volunteer, on_delete=models.CASCADE, related_name="combat_payments",
                                  verbose_name="Доброволец")
    date = models.DateField(verbose_name="Дата")
    amount = models.PositiveIntegerField(verbose_name="Сумма")

    class Meta:
        verbose_name = "Боевая выплата"
        verbose_name_plural = "Боевые выплаты"

    def __str__(self):
        return f"Боевая выплата {self.volunteer.last_name} {self.volunteer.first_name} - {self.amount} руб. ({self.date})"


class Remark(models.Model):
    volunteer = models.ForeignKey(Volunteer, on_delete=models.CASCADE, related_name="remarks",
                                  verbose_name="Доброволец")
    date = models.DateField(verbose_name="Дата нарекания")
    comment = models.TextField(blank=True, null=True, verbose_name="Комментарий к нареканию")

    class Meta:
        verbose_name = "Нарекание"
        verbose_name_plural = "Нарекания"

    def __str__(self):
        return f"Нарекание {self.volunteer.last_name} {self.volunteer.first_name} ({self.date})"


class Item(models.Model):
    CHARACTERISTICS_SCHEMA = {
        'type': 'list',
        'items': {
            'type': 'dict',
            'keys': {
                'name': {
                    'type': 'string',
                    'title': 'Название'
                },
                'value': {
                    'type': 'string',
                    'title': 'Значение'
                }
            }
        }
    }
    """Модель для предметов с характеристиками"""
    name = models.CharField(max_length=255, verbose_name="Название предмета")
    description = models.TextField(blank=True, null=True, verbose_name="Описание")
    characteristics = JSONField(blank=True, null=True, verbose_name='Характеристики', schema=CHARACTERISTICS_SCHEMA)

    class Meta:
        verbose_name = "Предмет"
        verbose_name_plural = "Предметы"

    def __str__(self):
        return self.name


class VolunteerItem(models.Model):
    """Связь между добровольцем и предметами"""
    volunteer = models.ForeignKey(Volunteer, on_delete=models.CASCADE, related_name="items", verbose_name="Доброволец")
    item = models.ForeignKey(Item, on_delete=models.CASCADE, related_name="volunteers", verbose_name="Предмет")
    quantity = models.PositiveIntegerField(verbose_name="Количество")

    class Meta:
        verbose_name = "Предмет для добровольца"
        verbose_name_plural = "Предметы для добровольцев"
        unique_together = (
            'volunteer',
            'item')

    def __str__(self):
        return f"{self.volunteer} - {self.item.name} (Кол-во: {self.quantity})"


class Report(models.Model):
    """Отчет за период"""
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="Дата создания отчета")
    file = models.FileField(upload_to="reports/", blank=True, null=True, verbose_name="Файл отчета")
    start_date = models.DateField(verbose_name="Дата начала периода", null=True)
    end_date = models.DateField(verbose_name="Дата конца периода", null=True)

    class Meta:
        verbose_name = "Отчет"
        verbose_name_plural = "Отчеты"

    def __str__(self):
        return f"Отчет с {self.start_date} по {self.end_date}"

    def generate_report(self):
        """Создание Excel-файла отчета"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет"

        headers = [
            "id", "№ Личный", "Статус", "Фамилия", "Имя", "Отчество",
            "Дата рождения", "Серия паспорта", "Номер паспорта", "Кем выдан паспорт",
            "Дата выдачи паспорта", "Дата контракта", "№ приказа", "Дата зачисления",
            "Размер денежной выплаты", "БИК", "Банк", "Корр. счет", "Расчетный счет", "ИНН", "КПП",
            "Кол-во отработанных дней"
        ]
        ws.append(headers)

        volunteers = get_volunteers_for_report(Volunteer.objects.all(), self.start_date, self.end_date)
        for volunteer in volunteers:
            row = [
                volunteer.id, volunteer.number_service, volunteer.get_status_display(),
                volunteer.last_name, volunteer.first_name, volunteer.patronymic,
                volunteer.birthday, volunteer.passport_series, volunteer.passport_number, volunteer.passport_issued,
                volunteer.passport_issue_date, volunteer.contract_date, volunteer.order_number,
                volunteer.enrollment_date, volunteer.salary_amount, volunteer.bic, volunteer.bank_name,
                volunteer.correspondent_account, volunteer.checking_account, volunteer.inn, volunteer.kpp,
                get_worked_days(volunteer, self.start_date, self.end_date)
            ]
            ws.append(row)

        file_stream = ContentFile(b"")
        wb.save(file_stream)
        file_stream.seek(0)
        filename = f"report_{self.start_date}_{self.end_date}.xlsx"
        self.file.save(filename, file_stream, save=False)

    def save(self, *args, **kwargs):
        self.generate_report()
        super().save(*args, **kwargs)

    @property
    def name(self):
        return self.__str__()

    name.fget.short_description = 'Название'


class ActivityReport(models.Model):
    """Отчет активности добровольцев"""
    STATUS_CHOICES = [
        ('pending', 'Ожидает обработки'),
        ('processing', 'В обработке'),
        ('completed', 'Завершено'),
        ('failed', 'Ошибка')
    ]
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="Дата создания отчета")
    report_date = models.DateField(verbose_name="Дата активности")
    file = models.FileField(upload_to="activity_reports/", verbose_name="Файл отчета")
    status = models.CharField(max_length=20, choices=STATUS_CHOICES, default='pending', verbose_name="Статус отчета")
    error_details = models.TextField(verbose_name="Детали ошибки", blank=True, null=True)  # Новое поле

    class Meta:
        verbose_name = "Отчет активности"
        verbose_name_plural = "Отчеты активности"

    def __str__(self):
        return f"Отчет активности ({self.report_date.strftime('%Y-%m-%d')})"

    def clean(self):
        """Валидация: файл и дата обязательны"""
        if not self.file:
            raise ValidationError({"file": _("Необходимо загрузить файл отчета.")})
        if not self.report_date:
            raise ValidationError({"report_date": _("Необходимо указать дату активности.")})

    def process_report(self):
        print(f"\n--- Начало обработки отчета от {self.report_date} ---")
        self.error_details = ""  # Сбрасываем предыдущие ошибки
        try:
            with transaction.atomic():
                # Загрузка файла
                print("[1/5] Загрузка файла Excel...")
                wb = openpyxl.load_workbook(self.file, data_only=True)
                ws = wb.active

                # Поиск столбца с табельными номерами
                print("[2/5] Поиск столбца с табельными номерами...")
                header_row = [str(cell.value).strip().lower() for cell in ws[1]]
                tn_col_index = next(
                    (idx for idx, h in enumerate(header_row)
                     if any(kw in h for kw in {"Личный номер", "личный номер"})),
                    None
                )

                if tn_col_index is None:
                    self.error_details = "❌ Столбец с табельными номерами не найден!"
                    self.status = 'failed'
                    self.save(update_fields=['status', 'error_details'])
                    raise ValueError(self.error_details)

                print(f"✔️ Табельные номера в столбце {chr(65 + tn_col_index)}")

                # Сбор данных из файла
                print("[3/5] Чтение данных из файла...")
                report_tn = set()
                invalid_tn = []
                for row_num, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
                    tn = str(row[tn_col_index]).strip()
                    if not tn:
                        continue

                    report_tn.add(tn)

                if invalid_tn:
                    self.error_details = "❌ Некорректные табельные номера:\n" + "\n".join(invalid_tn)
                    self.status = 'failed'
                    self.save(update_fields=['status', 'error_details'])
                    raise ValueError(self.error_details)

                print(f"Найдено табельных номеров в отчете: {len(report_tn)}")

                # Увольнение отсутствующих волонтеров
                print("[4/5] Проверка активных волонтеров...")
                active_volunteers = Volunteer.objects.filter(status='active')
                dismissed = []

                for volunteer in active_volunteers:
                    if volunteer.number_service not in report_tn:
                        volunteer.status = 'dismissed'
                        volunteer.dismissal_date = self.report_date
                        volunteer.dismissal_order_number = f"Автоувольнение {self.report_date}"
                        dismissed.append(volunteer)

                if dismissed:
                    Volunteer.objects.bulk_update(dismissed, ['status', 'dismissal_date', 'dismissal_order_number'])
                    print(f"🚫 Уволено волонтеров: {len(dismissed)}")
                else:
                    print("🤷 Нет волонтеров для увольнения")

                # Добавление новых волонтеров
                print("[5/5] Обработка новых волонтеров...")
                existing_numbers = set(Volunteer.objects.values_list('number_service', flat=True))
                new_numbers = report_tn - existing_numbers

                new_volunteers = []
                errors = []
                if new_numbers:
                    for row_num, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
                        tn = row[tn_col_index]
                        if tn is None or str(tn).strip() == "":  # Проверяем None и пустую строку
                            continue
                        if tn in new_numbers:
                            try:
                                print(f"\n--- Строка {row_num} ---")
                                print(f"ТН: {tn}")
                                print(f"Фамилия (row[10]): {row[10]}")
                                print(f"Имя (row[11]): {row[11]}")
                                print(f"Отчество (row[12]): {row[12]}")
                                print(f"Дата рождения (row[14]): {row[14]}")
                                print(f"Паспорт (серия: row[36]={row[36]}, номер: row[37]={row[37]})")
                                print(f"Кем выдан (row[39]): {row[39]}")
                                print(f"Дата выдачи паспорта (row[38]): {row[38]}")
                                print(f"Дата договора (row[82]): {row[82]}")
                                print(f"Номер приказа (row[81]): {row[81]}")
                                print(f"Дата зачисления (row[85]): {row[85]}")
                                print(f"БИК (row[52]): {row[52]}")
                                print(f"ИНН (row[48]): {row[48]}")
                                print(f"Расчетный счет (row[53]): {row[53]}")
                                print(f"Звание (row[6]): {row[6]}")

                                # Парсинг дат с обработкой ошибок
                                def parse_date(date_str, field_name):
                                    if not date_str:
                                        return None
                                    if isinstance(date_str, datetime):  # Если уже datetime, преобразуем в дату
                                        return date_str.date()
                                    if isinstance(date_str, date):  # Если уже date, возвращаем как есть
                                        return date_str
                                    try:
                                        return datetime.strptime(date_str, "%d.%m.%Y").date()
                                    except ValueError:
                                        errors.append(
                                            f"Строка {row_num}: Некорректный формат {field_name} ('{date_str}')")
                                        return None

                                birthday = parse_date(row[14], "даты рождения")
                                passport_issue_date = parse_date(row[38], "даты выдачи паспорта")
                                contract_date = parse_date(row[82], "даты договора")
                                enrollment_date = parse_date(row[85], "даты зачисления")

                                # Проверяем наличие критических ошибок
                                if errors and len(errors) >= 5:  # Пример ограничения
                                    raise ValueError("Слишком много ошибок в данных")

                                new_volunteers.append(Volunteer(
                                    number_service=tn,
                                    last_name=row[10],
                                    first_name=row[11],
                                    patronymic=row[12],
                                    birthday=birthday,
                                    passport_series=row[36],
                                    passport_number=row[37],
                                    passport_issued=row[39],
                                    passport_issue_date=passport_issue_date,
                                    contract_date=contract_date,
                                    order_number=row[81],
                                    enrollment_date=enrollment_date,
                                    bic=row[52],
                                    inn=row[48],
                                    checking_account=row[53],
                                    rank=row[6],
                                    status='active',
                                    salary_amount=0.00
                                ))
                            except Exception as e:
                                errors.append(f"Строка {row_num}: {str(e)}")
                                continue

                    if errors:
                        self.error_details = "❌ Ошибки при создании волонтеров:\n" + "\n".join(errors)
                        self.status = 'failed'
                        self.save(update_fields=['status', 'error_details'])
                        raise ValueError(self.error_details)

                    if new_volunteers:
                        Volunteer.objects.bulk_create(new_volunteers)
                        print(f"✅ Добавлено новых волонтеров: {len(new_volunteers)}")
                else:
                    print("🤷 Нет новых волонтеров для добавления")

                # Если все успешно
                self.status = 'completed'
                self.error_details = ""

        except Exception as e:
            print(f"🔥 Критическая ошибка: {str(e)}")
            self.status = 'failed'
            if not self.error_details:  # Если ошибка не была записана ранее
                self.error_details = str(e)
        finally:
            # Сохраняем статус и ошибки в отдельной транзакции
            try:
                with transaction.atomic():
                    self.save(update_fields=['status', 'error_details'])
            except Exception as e:
                print(f"Ошибка при сохранении статуса: {e}")

        print("--- Обработка отчета завершена ---\n")

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        if self.status == 'pending':  # Запускаем обработку только при создании
            self.process_report()


class UpdateReport(models.Model):
    """Отчет обновления данных добровольцев"""
    STATUS_CHOICES = [
        ('pending', 'Ожидает обработки'),
        ('processing', 'В обработке'),
        ('completed', 'Завершено'),
        ('failed', 'Ошибка')
    ]

    created_at = models.DateTimeField(auto_now_add=True, verbose_name="Дата создания отчета")
    file = models.FileField(upload_to="update_reports/", verbose_name="Файл отчета")
    status = models.CharField(max_length=20, choices=STATUS_CHOICES, default='pending', verbose_name="Статус отчета")
    error_details = models.TextField(verbose_name="Детали ошибки", blank=True, null=True)

    class Meta:
        verbose_name = "Отчет обновления"
        verbose_name_plural = "Отчеты обновления"

    def __str__(self):
        return f"Обновление данных ({self.created_at.strftime('%Y-%m-%d')})"

    def clean(self):
        if not self.file:
            raise ValidationError({"file": _("Необходимо загрузить файл отчета.")})

    def process_report(self):
        print(f"\n--- Начало обработки отчета обновления данных ---")
        self.error_details = ""  # Сбрасываем предыдущие ошибки
        try:
            with transaction.atomic():
                print("[1/4] Загрузка файла Excel...")
                wb = openpyxl.load_workbook(self.file, data_only=True)
                ws = wb.active

                print("[2/4] Чтение заголовков...")
                header_row = [str(cell.value).strip().lower() for cell in ws[3]]
                column_map = {
                    'number_service': next((idx for idx, h in enumerate(header_row) if 'личный номер' in h), None),
                    'last_name': next((idx for idx, h in enumerate(header_row) if 'фамилия' in h), None),
                    'first_name': next((idx for idx, h in enumerate(header_row) if 'имя' in h), None),
                    'patronymic': next((idx for idx, h in enumerate(header_row) if 'отчество' in h), None),
                    'birthday': next((idx for idx, h in enumerate(header_row) if 'дата рождения' in h), None),
                    'bic': next((idx for idx, h in enumerate(header_row) if 'бик' in h), None),
                    'correspondent_account': next((idx for idx, h in enumerate(header_row) if 'номер счета' in h),
                                                  None),
                }

                if None in column_map.values():
                    self.error_details = "❌ Один или несколько обязательных столбцов не найдены!"
                    self.status = 'failed'
                    self.save(update_fields=['status', 'error_details'])
                    raise ValueError(self.error_details)

                print("✔️ Заголовки успешно считаны")

                print("[3/4] Обновление данных волонтеров...")
                updated_volunteers = []
                errors = []

                for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                    number_service = row[column_map['number_service']]
                    if not number_service:
                        continue

                    try:
                        volunteer = Volunteer.objects.get(number_service=number_service)
                        volunteer.last_name = row[column_map['last_name']]
                        volunteer.first_name = row[column_map['first_name']]
                        volunteer.patronymic = row[column_map['patronymic']]
                        volunteer.birthday = row[column_map['birthday']]
                        volunteer.bic = row[column_map['bic']]
                        volunteer.correspondent_account = row[column_map['correspondent_account']]
                        updated_volunteers.append(volunteer)
                    except Volunteer.DoesNotExist:
                        errors.append(f"Строка {row_num}: Волонтер с номером {number_service} не найден.")

                if errors:
                    self.error_details = "❌ Ошибки при обновлении волонтеров:\n" + "\n".join(errors)
                    self.status = 'failed'
                    self.save(update_fields=['status', 'error_details'])
                    raise ValueError(self.error_details)

                if updated_volunteers:
                    Volunteer.objects.bulk_update(updated_volunteers,
                                                  ['last_name', 'first_name', 'patronymic', 'birthday', 'bic',
                                                   'correspondent_account'])
                    print(f"✅ Обновлено волонтеров: {len(updated_volunteers)}")
                else:
                    print("🤷 Нет данных для обновления")

                self.status = 'completed'
                self.error_details = ""

        except Exception as e:
            print(f"🔥 Критическая ошибка: {str(e)}")
            self.status = 'failed'
            if not self.error_details:
                self.error_details = str(e)
        finally:
            try:
                with transaction.atomic():
                    self.save(update_fields=['status', 'error_details'])
            except Exception as e:
                print(f"Ошибка при сохранении статуса: {e}")

        print("--- Обработка отчета завершена ---\n")

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        if self.status == 'pending':
            self.process_report()


class SalaryReport(models.Model):
    """Отчет о зарплате волонтеров за период"""
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="Дата создания отчета")
    start_date = models.DateField(verbose_name="Дата начала периода")
    end_date = models.DateField(verbose_name="Дата окончания периода")
    file = models.FileField(upload_to="salary_reports/", blank=True, null=True, verbose_name="Файл отчета")

    class Meta:
        verbose_name = "Расчетный лист"
        verbose_name_plural = "Расчетные листы"

    def __str__(self):
        return f"Расчетный лист с {self.start_date} по {self.end_date}"

    def clean(self):
        """Валидация дат"""
        if self.start_date >= self.end_date:
            raise ValidationError({"end_date": "Дата окончания должна быть позже даты начала."})

    def generate_report(self):
        """Генерация Excel-файла с расчетным листом"""

        volunteers = get_volunteers_for_report(Volunteer.objects.all(), self.start_date, self.end_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Расчетный лист"

        headers = ["ФИО", "Личный номер", "Должность", "Оклад", "Боевые", "Губернаторские выплаты", "Итого"]
        ws.append(headers)

        # Вычисляем количество месяцев в периоде
        num_months = (self.end_date.year - self.start_date.year) * 12 + self.end_date.month - self.start_date.month + 1

        for volunteer in volunteers:
            full_name = f"{volunteer.last_name} {volunteer.first_name} {volunteer.patronymic or ''}".strip()
            rank = volunteer.rank or "—"
            salary = volunteer.salary * num_months  # Оклад за период

            # Считаем боевые выплаты за период
            combat_total = Combat.objects.filter(
                volunteer=volunteer,
                date__range=(self.start_date, self.end_date)
            ).aggregate(total=models.Sum("amount"))["total"] or 0

            # Считаем губернаторские выплаты
            worked_days = get_worked_days(volunteer, self.start_date, self.end_date)
            governor_payments = worked_days * 1457

            # Итоговая сумма
            total_amount = salary + combat_total + governor_payments

            # Добавляем строку в Excel
            ws.append(
                [full_name, volunteer.number_service, rank, salary, combat_total, governor_payments, total_amount])

        # Сохраняем файл в `self.file`
        file_stream = ContentFile(b"")
        wb.save(file_stream)
        file_stream.seek(0)
        filename = f"salary_report_{self.start_date}_{self.end_date}.xlsx"
        self.file.save(filename, file_stream, save=False)

    def save(self, *args, **kwargs):
        """Перед сохранением создаем отчет"""
        self.full_clean()
        super().save(*args, **kwargs)
        self.generate_report()
        super().save(update_fields=["file"])
