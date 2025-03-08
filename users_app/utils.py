import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment


def export_to_excel(queryset, filename):
    """Экспортирует данные в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteers"

    is_dismissed = queryset.first().status == "dismissed" if queryset.exists() else False

    headers = [
        "id", "№ личный", "Статус", "Фамилия", "Имя", "Отчество",
        "Дата рождения", "Серия паспорта", "Номер паспорта", "Кем выдан паспорт",
        "Дата выдачи паспорта", "Дата контракта", "№ приказа", "Дата зачисления",
        "Размер денежной выплаты", "БИК", "Банк", "Корр. счет", "Расчетный счет", "ИНН", "КПП"
    ]

    if is_dismissed:
        headers.extend(["Дата увольнения", "№ приказа об увольнении"])

    ws.append(headers)

    for volunteer in queryset:
        row = [
            volunteer.id, volunteer.number_service, volunteer.get_status_display(),
            volunteer.last_name, volunteer.first_name, volunteer.patronymic,
            volunteer.birthday, volunteer.passport_series, volunteer.passport_number, volunteer.passport_issued,
            volunteer.passport_issue_date, volunteer.contract_date, volunteer.order_number,
            volunteer.enrollment_date, volunteer.salary_amount, volunteer.bic, volunteer.bank_name,
            volunteer.correspondent_account, volunteer.checking_account, volunteer.inn, volunteer.kpp
        ]

        if is_dismissed:
            row.extend([volunteer.dismissal_date, volunteer.dismissal_order_number])

        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_volunteers_and_items_to_excel(queryset, filename):
    """Экспортирует данные добровольцев и связанных с ними предметов в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteers and Items"

    headers = [
        "id", "№ личный", "Статус", "Фамилия", "Имя", "Отчество",
        "Дата рождения", "Серия паспорта", "Номер паспорта", "Кем выдан паспорт",
        "Дата выдачи паспорта", "Дата контракта", "№ приказа", "Дата зачисления",
        "Размер денежной выплаты", "БИК", "Банк", "Корр. счет", "Расчетный счет", "ИНН", "КПП",
        "Предмет", "Описание предмета", "Характеристики", "Количество"
    ]
    ws.append(headers)

    row_start = 2

    for volunteer in queryset:
        volunteer_data = [
            volunteer.id, volunteer.number_service, volunteer.get_status_display(),
            volunteer.last_name, volunteer.first_name, volunteer.patronymic,
            volunteer.birthday, volunteer.passport_series, volunteer.passport_number, volunteer.passport_issued,
            volunteer.passport_issue_date, volunteer.contract_date, volunteer.order_number,
            volunteer.enrollment_date, volunteer.salary_amount, volunteer.bic, volunteer.bank_name,
            volunteer.correspondent_account, volunteer.checking_account, volunteer.inn, volunteer.kpp
        ]

        items = volunteer.items.all()

        if items:
            for i, item_relation in enumerate(items):
                item = item_relation.item
                characteristics = item.characteristics if item.characteristics else []

                characteristics_str = ", ".join([f"{char['name']}:{char['value']}" for char in characteristics])

                item_data = [
                    item.name,
                    item.description,
                    characteristics_str,
                    item_relation.quantity
                ]
                ws.append(volunteer_data + item_data)

            for col in range(1, len(volunteer_data) + 1):
                ws.merge_cells(
                    start_row=row_start,
                    start_column=col,
                    end_row=row_start + len(items) - 1,
                    end_column=col
                )

            row_start += len(items)
        else:
            ws.append(volunteer_data)
            row_start += 1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
